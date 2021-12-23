import openpyxl as xl
absolutely = ''


def process_excel(filename):
    # wb = xl.load_workbook('E:\Dev\project\mypython\myexcel.xlsx')
    wb = xl.load_workbook(filename)
    sheet1 = wb['Sheet1']
    cell = sheet1['a1']
    cell = sheet1.cell(1, 1)
    # print(cell.value)
    # print(sheet1.max_row)

    # assert isinstance(wb, xl.workbook)

    cell2 = sheet1.cell(1, 5)

    for row in range(2, sheet1.max_row + 1):
        cell = sheet1.cell(row, 1)
        newValue = cell.value * 0.9
        cell4 = sheet1.cell(row, 4)
        cell4.value = newValue

    # 导入图标包
    from openpyxl.chart import BarChart, Reference

    # 用来选中单元格的函数
    values = Reference(sheet1,
                       min_row=2,
                       max_row=4,
                       min_col=3,
                       max_col=3)

    # 新建一个条形图
    chart = BarChart();
    chart.add_data(values)
    # 先增加一个图表，然后是图表的坐标，首先是左上角的坐标
    sheet1.add_chart(chart, 'e2')

    # 保存到新的文件
    # wb.save('newExcel.xlsx')
    # 重写原来的文件
    wb.save(filename)


